import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.hyperlink import Hyperlink

# Read the Excel file
df = pd.read_excel("buy-and-sell_facebook_groups_dataset.xlsx")

# Drop duplicate rows based on "Group URL" column
df_unique = df.drop_duplicates(subset=["Group URL"])
# Replace '0' in "Today's Post" column with "No New Post Today"
df_unique["Today's Post"] = df_unique["Today's Post"].apply(lambda x: "No New Post Today" if x == 0 else x)

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Add column headers to the worksheet
headers = list(df_unique.columns)
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)

# Add data from DataFrame to worksheet
for r_idx, row in enumerate(dataframe_to_rows(df_unique, index=False, header=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Add hyperlinks
for row in ws.iter_rows(min_row=2, min_col=headers.index("Group Name") + 1, max_col=headers.index("Group Name") + 1, max_row=len(df_unique) + 1):
    for cell in row:
        group_name = cell.value
        group_url = df_unique.loc[df_unique["Group Name"] == group_name, "Group URL"].iloc[0]
        cell.hyperlink = Hyperlink(ref=group_url, target=group_url)
        # Apply formatting to resemble a hyperlink
        cell.font = Font(color="0000FF", underline="single")

# Save the workbook
wb.save("buy-and-sell_facebook_groups_dataset_hyperlink.xlsx")
